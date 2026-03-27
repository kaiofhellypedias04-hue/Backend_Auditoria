# modules/nfse_xml_converter.py
"""Módulo para conversão de XMLs NFS-e com auditoria fiscal (modo ALERTA)

Fluxo (atualizado):
1) Extrai os dados dos XMLs e gera a planilha mantendo 100% dos valores informados no XML (sem sobrescrever).
2) Usa a tabela de regras (LC116/retencões) apenas para VALIDAR e GERAR "Alertas Fiscais":
   - Se o código de serviço estiver na tabela, o sistema compara o que veio no XML vs o que seria esperado.
   - Se faltar imposto esperado ou houver imposto "a mais", registra em "Alertas Fiscais".
3) Camada adicional: quando o XML indicar Simples Nacional ou MEI, as retenções federais (IRRF/CSRF) são tratadas
   como potencialmente indevidas e também geram alertas (sem alterar valores).

Observação:
- Este módulo não "corrige" valores do XML. Ele apenas sinaliza divergências.
"""
import xml.etree.ElementTree as ET
import re
import unicodedata
import pandas as pd
from .fiscal_status import compute_base_calculation_status, compute_final_note_status
import os
import traceback
from datetime import datetime
from .cnpj_consultor import CNPJConsultor
from .cache import obter_estatisticas_cache

class NFSeXMLConverter:
    """Classe principal para conversão de XMLs NFS-e com auditoria fiscal"""
    
    def __init__(self, tipo_nota: str = "tomados"):
        """
        Inicializa o conversor de XML NFS-e.
        
        Args:
            tipo_nota: 'tomados' (Recebidas) ou 'prestados' (Emitidas)
                       - tomados: usa CNPJ do emitente (prestador de serviço)
                       - prestados: usa CNPJ do tomador (cliente que recebeu o serviço)
        """
        self.namespaces = {
            'ns': 'http://www.sped.fazenda.gov.br/nfse',
            'ds': 'http://www.w3.org/2000/09/xmldsig#'
        }
        self.tipo_nota = tipo_nota
        
                # =====================================================================
        # REGRAS DE RETENÇÃO (IRRF / CSRF / INSS) - FONTE: PLANILHA "RETENÇÕES - REGRAS"
        # - Usadas SOMENTE para VALIDAR e GERAR ALERTAS (não altera valores do XML).
        # - Qualquer regra hardcoded foi removida. A planilha é a única fonte.
        # =====================================================================
        self.REGRAS_RETENCOES = self._carregar_regras_planilha()
        print(f"📘 Regras (planilha) carregadas: {len(self.REGRAS_RETENCOES)} códigos")
        
        # ------------------------------------------------------------
        # CÓDIGOS QUE SEMPRE DEVEM TER RETENÇÃO DE INSS
        # (independente de cessão ou regime)
        # ------------------------------------------------------------
        self.INSS_CODIGOS_OBRIGATORIOS = ["7.02", "11.02"]
    
    # ------------------------------------------------------------------
    # NORMALIZAÇÃO DE ALÍQUOTAS (percentual -> decimal)
    # ------------------------------------------------------------------
    def _normalizar_codigo_servico(self, codigo: str):
        """Normaliza códigos como '07.02', '7.2', '7.02' -> '7.02'. Retorna None se inválido."""
        if not codigo:
            return None
        m = re.search(r"(\d{1,2})\s*\.\s*(\d{2})", str(codigo))
        if not m:
            return None
        return f"{int(m.group(1))}.{m.group(2)}"

    def _carregar_regras_planilha(self):
        """
        Carrega regras de retenção da planilha embutida no projeto.
        Arquivo: test_planilha/data/RETENCOES_REGRAS.xlsx (aba 'Planilha2')
        Colunas esperadas (Planilha2):
          B: Subitem (ex: 7.02)
          D: IRRF (SIM/NÃO/DEPENDE)
          E: Alíquota (ex: 0.015)
          F: CSRF (SIM/NÃO/DEPENDE)
          G: INSS (SIM/NÃO/DEPENDE)
          H: Observação
        """
        try:
            import openpyxl
            base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
            xlsx_path = os.path.join(base_dir, "data", "RETENCOES_REGRAS.xlsx")

            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            ws = wb["Planilha2"] if "Planilha2" in wb.sheetnames else wb.active

            regras = {}
            for row in range(2, ws.max_row + 1):
                sub = ws.cell(row, 2).value  # coluna B
                codigo = self._normalizar_codigo_servico(sub)
                if not codigo:
                    continue

                irrf_flag = str(ws.cell(row, 4).value or "").strip().upper()   # D
                aliq = ws.cell(row, 5).value                                   # E
                csrf_flag = str(ws.cell(row, 6).value or "").strip().upper()   # F
                inss_flag = str(ws.cell(row, 7).value or "").strip().upper()   # G
                obs = ws.cell(row, 8).value                                    # H

                # IRRF: se SIM e tiver aliquota -> percent
                irrf_percent = None
                if irrf_flag in ("NÃO", "NAO"):
                    irrf_percent = 0.0
                elif irrf_flag == "SIM":
                    if aliq is not None and str(aliq).strip() != "":
                        try:
                            irrf_percent = float(aliq) * 100.0  # 0.015 -> 1.5
                        except Exception:
                            irrf_percent = None

                # CSRF: SIM -> 4.65, NÃO -> 0.0
                csrf_percent = None
                if csrf_flag in ("NÃO", "NAO"):
                    csrf_percent = 0.0
                elif csrf_flag == "SIM":
                    csrf_percent = 4.65

                # INSS: SIM/NÃO/DEPENDE (bool ou None)
                inss_bool = None
                if inss_flag in ("NÃO", "NAO"):
                    inss_bool = False
                elif inss_flag == "SIM":
                    inss_bool = True

                regras[codigo] = {
                    "irrf_flag": irrf_flag,
                    "irrf_percent": irrf_percent,
                    "csrf_flag": csrf_flag,
                    "csrf_percent": csrf_percent,
                    "inss_flag": inss_flag,
                    "inss_bool": inss_bool,
                    "obs": obs,
                }

            return regras
        except Exception as e:
            print(f"⚠️ Falha ao carregar planilha de regras: {e}")
            return {}

    def _normalize_rate(self, percent):
        """Converte percentual (ex: 1.5) para decimal (0.015)."""
        try:
            p = float(percent)
            if p >= 1.0:
                return p / 100.0
            else:
                return p
        except:
            return 0.0

    # ------------------------------------------------------------------
    # NORMALIZAÇÃO DO CÓDIGO DE SERVIÇO
    # ------------------------------------------------------------------
    def normalizar_codigo_servico(self, codigo):
        if not codigo or '.' not in codigo:
            return codigo
        parte1, parte2 = codigo.split('.', 1)
        parte1 = parte1.lstrip('0')
        if parte1 == '':
            parte1 = '0'
        return f"{parte1}.{parte2}"
    
    # ------------------------------------------------------------------
    # EXTRAÇÃO ROBUSTA DE CÓDIGO DE SERVIÇO (3,4,5,6 DÍGITOS)
    # ------------------------------------------------------------------
    def _extrair_codigo_lc116(self, codigo_raw):
        if not codigo_raw:
            return ''
        numeros = re.sub(r'\D', '', codigo_raw)
        
        if len(numeros) == 3:
            return f"{numeros[0]}.{numeros[1:]}"
        elif len(numeros) == 4:
            return f"{int(numeros[:2])}.{numeros[2:]}"
        elif len(numeros) == 5:
            return f"{int(numeros[:2])}.{numeros[2:4]}"
        elif len(numeros) == 6:
            return f"{int(numeros[:2])}.{numeros[2:4]}"
        else:
            return codigo_raw
    
    # ------------------------------------------------------------------
    # Métodos auxiliares (conversão, formatação)
    # ------------------------------------------------------------------
    def _to_float(self, value):
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return 0.0
            if ',' in value:
                value = value.replace('.', '').replace(',', '.')
        try:
            return float(value)
        except:
            return 0.0
    
    def format_cnpj(self, cnpj):
        if not cnpj:
            return ''
        cnpj = str(cnpj).strip().replace('.', '').replace('/', '').replace('-', '')
        if len(cnpj) == 14:
            return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
        elif len(cnpj) == 11:
            return f"{cnpj[:3]}.{cnpj[3:6]}.{cnpj[6:9]}-{cnpj[9:]}"
        return cnpj
    
    def format_currency(self, value):
        try:
            if not value or value == '':
                return '0.00'
            if isinstance(value, str):
                value = self._to_float(value)
            return f"{value:.2f}"
        except:
            return '0.00'
    # ------------------------------------------------------------------
    # Regras (Planilha) - modo ALERTA (NÃO força nada)
    # ------------------------------------------------------------------

    def aplicar_regras_retencao(self, dados, codigo_servico):
        """(Modo ALERTA) Calcula IRRF/CSRF esperados conforme a planilha.
        Não altera 'dados'. Retorna dict com valores esperados quando aplicável.
        """
        valor_total = self._to_float(dados.get("Valor Total", 0))
        valor_bc = self._to_float(dados.get("Valor B/C", 0))
        base_calculo = valor_bc if valor_bc > 0 else valor_total

        codigo_norm = self._normalizar_codigo_servico(codigo_servico) or codigo_servico
        regra = self.REGRAS_RETENCOES.get(codigo_norm)
        if not regra:
            return {"status": "SEM_REGRA", "codigo": codigo_norm}

        irrf_percentual = regra.get("irrf_percent")
        csrf_percentual = regra.get("csrf_percent")

        out = {"status": "OK", "codigo": codigo_norm, "obs": regra.get("obs")}
        if irrf_percentual is not None:
            out["irrf_percentual"] = irrf_percentual
            out["irrf_esperado"] = base_calculo * (float(irrf_percentual) / 100.0)
        if csrf_percentual is not None:
            out["csrf_percentual"] = csrf_percentual
            out["csrf_esperado"] = base_calculo * (float(csrf_percentual) / 100.0)
        return out

    def aplicar_regras_inss(self, dados, codigo_servico):
        """(Modo ALERTA) Lê a expectativa INSS (SIM/NÃO/DEPENDE) pela planilha.
        Não altera 'dados'.
        """
        codigo_norm = self._normalizar_codigo_servico(codigo_servico) or codigo_servico
        regra = self.REGRAS_RETENCOES.get(codigo_norm)
        if not regra:
            return {"status": "SEM_REGRA", "codigo": codigo_norm}

        return {
            "status": "OK",
            "codigo": codigo_norm,
            "inss_flag": regra.get("inss_flag"),
            "inss_bool": regra.get("inss_bool"),
            "obs": regra.get("obs"),
        }

    def verificar_correcoes(self, dados):
        """Gera texto de alertas fiscais (modo ALERTA).
        Só valida se o código existir na planilha. Não altera valores.
        """
        codigo_servico = dados.get('Código de serviço', '') or ''
        codigo_norm = self._normalizar_codigo_servico(codigo_servico) or codigo_servico
        if codigo_norm not in self.REGRAS_RETENCOES:
            return ''

        # Usa 'Simples Nacional / XML' como fonte do regime (não 'Regime')
        regime = dados.get('Simples Nacional / XML', '') or ''
        tipo_retencao_csrf = dados.get('Tipo retenção CSRF', '') or ''
        valor_irrf = self._to_float(dados.get('IRRF', 0))
        valor_inss = self._to_float(dados.get('INSS', 0))
        tem_irrf_retido = bool(valor_irrf > 0)
        tem_inss_retido = bool(valor_inss > 0)

        validacao = self.validar_retencoes(
            dados, regime, codigo_norm, tipo_retencao_csrf, tem_irrf_retido, tem_inss_retido
        )
        alertas = validacao.get('alertas', []) if isinstance(validacao, dict) else []
        return " | ".join(alertas) if alertas else ''



    def validar_retencoes(self, dados, regime, codigo_servico, tipo_retencao_csrf, tem_irrf_retido, tem_inss_retido):
        """
        Valida as retenções com base APENAS na planilha de regras (Planilha2).
        Gera alertas, mas NÃO altera os valores.
        Só valida quando o código de serviço existir na planilha.
        
        REGRAS:
        - MEI: SEMPRE sem retenção (não valida mais nada)
        - Optante do Simples (não MEI): Sem IRRF/CSRF, mas segue Anexo IV para INSS
        - Demais empresas: Segue regras da planilha (IRRF, CSRF, INSS)
        - BASE ZERADA: Sempre gera divergência se tiver imposto retido
        """
        alertas = []
        correcoes = {}

        # normaliza código e checa se existe na planilha
        codigo_norm = self._normalizar_codigo_servico(codigo_servico) or codigo_servico
        regra = self.REGRAS_RETENCOES.get(codigo_norm)
        if not regra:
            return {'alertas': alertas, 'correcoes': correcoes}

        valor_total = self._to_float(dados.get('Valor Total', 0))
        valor_bc = self._to_float(dados.get('Valor B/C', 0))
        valor_irrf = self._to_float(dados.get('IRRF', 0))
        valor_csrf = self._to_float(dados.get('CSRF', 0))
        valor_inss = self._to_float(dados.get('INSS', 0))

        base_calculo = valor_bc if valor_bc > 0 else valor_total

        # ================================================================
        # REGRA 0: BASE ZERADA - Sempre gera divergência se tiver imposto retido
        # (Aplica-se a TODOS os regimes)
        # ================================================================
        base_zerada = (valor_bc == 0 and valor_total > 0)
        
        if base_zerada:
            # Base zerada com IRRF retido = divergência
            if valor_irrf > 0:
                alertas.append(
                    f"BASE ZERADA: IRRF retido ({valor_irrf:.2f}) mas base de cálculo é zero. Verificar!"
                )
            # Base zerada com CSRF retido = divergência
            if valor_csrf > 0:
                alertas.append(
                    f"BASE ZERADA: CSRF retido ({valor_csrf:.2f}) mas base de cálculo é zero. Verificar!"
                )
            # Base zerada com INSS retido = divergência
            if valor_inss > 0:
                alertas.append(
                    f"BASE ZERADA: INSS retido ({valor_inss:.2f}) mas base de cálculo é zero. Verificar!"
                )

        # Helpers esperados
        exp_irrf_percent = regra.get('irrf_percent')  # float em %
        exp_irrf_flag = regra.get('irrf_flag')        # SIM/NÃO/DEPENDE
        exp_csrf_percent = regra.get('csrf_percent')  # float em %
        exp_csrf_flag = regra.get('csrf_flag')        # SIM/NÃO/DEPENDE
        exp_inss_bool = regra.get('inss_bool')        # True/False/None
        exp_inss_flag = regra.get('inss_flag')        # SIM/NÃO/DEPENDE

        # ================================================================
        # REGRA 1: MEI - SEMPRE sem retenção (não valida mais nada)
        # ================================================================
        if regime == 'MEI':
            if valor_irrf != 0 or tem_irrf_retido:
                alertas.append(
                    f"MEI: IRRF não deve ser retido. Valor encontrado: {valor_irrf:.2f}"
                )
            if valor_csrf != 0 or (tipo_retencao_csrf and str(tipo_retencao_csrf).strip()):
                alertas.append(
                    f"MEI: CSRF não deve ser retido. Valor encontrado: {valor_csrf:.2f}"
                )
            if valor_inss != 0:
                alertas.append(
                    f"MEI: INSS não deve ser retido. Valor encontrado: {valor_inss:.2f}"
                )
            # Para MEI, não validamos mais nada - retorna aqui
            return {'alertas': alertas, 'correcoes': correcoes}

        # ================================================================
        # REGRA 2: OPTANTE DO SIMPLES (não MEI) - Sem IRRF/CSRF, segue Anexo IV para INSS
        # IMPORTANTE: Mesmo com base zerada, se tiver IRRF/CSRF retido é DIVERGENTE
        # ================================================================
        if regime in ['Optante S.N', 'Simples Nacional', 'Simples']:
            # IRRF: SEMPRE sem retenção para optantes (mesmo com base zerada)
            if valor_irrf != 0 or tem_irrf_retido:
                alertas.append(
                    f"Optante Simples: IRRF não deve ser retido para código {codigo_norm}. Valor encontrado: {valor_irrf:.2f}"
                )
            # CSRF: SEMPRE sem retenção para optantes (mesmo com base zerada)
            if valor_csrf != 0 or (tipo_retencao_csrf and str(tipo_retencao_csrf).strip()):
                alertas.append(
                    f"Optante Simples: CSRF não deve ser retido para código {codigo_norm}. Valor encontrado: {valor_csrf:.2f}"
                )
            # Se base zerada e tem algum imposto retido, já gerou alerta acima
            # Agora valida INSS conforme Anexo IV (só faz sentido se base > 0)
            if not base_zerada and base_calculo > 0:
                if exp_inss_bool is True and valor_inss <= 0:
                    alertas.append(
                        f"Optante Simples: INSS esperado (Anexo IV=SIM) para código {codigo_norm}, mas veio 0.00."
                    )
                elif exp_inss_bool is False and valor_inss > 0:
                    alertas.append(
                        f"Optante Simples: INSS não esperado (Anexo IV=NÃO) para código {codigo_norm}, mas veio {valor_inss:.2f}."
                    )
                elif exp_inss_flag == 'DEPENDE' and valor_inss > 0:
                    alertas.append(
                        f"Optante Simples: INSS na planilha = DEPENDE para código {codigo_norm}. Revisar regra/observação."
                    )
            # Para optantes, não validamos IRRF/CSRF da planilha - retorna aqui
            return {'alertas': alertas, 'correcoes': correcoes}

        # ================================================================
        # REGRA 3: DEMAIS EMPRESAS (Não Simples/MEI) - Segue regras da planilha
        # ================================================================
        
        # IRRF
        if exp_irrf_percent is not None:
            valor_esperado = base_calculo * (float(exp_irrf_percent) / 100.0)
            if float(exp_irrf_percent) == 0.0:
                if valor_irrf != 0 or tem_irrf_retido:
                    alertas.append(
                        f"IRRF não deve ser retido para código {codigo_norm}. Valor encontrado: {valor_irrf:.2f}"
                    )
            else:
                # se tem retenção, compara valor; se não tem e esperado>0, alerta
                if valor_irrf > 0 or tem_irrf_retido:
                    if abs(valor_irrf - valor_esperado) > 0.01:
                        alertas.append(
                            f"IRRF retido divergente para código {codigo_norm}. Esperado: {valor_esperado:.2f} ({exp_irrf_percent}%). Encontrado: {valor_irrf:.2f}"
                        )
                else:
                    if valor_esperado > 0.01:
                        alertas.append(
                            f"IRRF devido e não retido para código {codigo_norm}. Deveria ser: {valor_esperado:.2f} ({exp_irrf_percent}%)"
                        )
        elif exp_irrf_flag == 'DEPENDE' and (valor_irrf > 0 or tem_irrf_retido):
            alertas.append(
                f"IRRF na planilha = DEPENDE para código {codigo_norm}. Revisar regra/observação."
            )

        # CSRF
        if exp_csrf_percent is not None:
            valor_esperado = base_calculo * (float(exp_csrf_percent) / 100.0)
            if float(exp_csrf_percent) == 0.0:
                if valor_csrf != 0 or (tipo_retencao_csrf and str(tipo_retencao_csrf).strip()):
                    alertas.append(
                        f"CSRF não deve ser retido para código {codigo_norm}. Valor encontrado: {valor_csrf:.2f}"
                    )
            else:
                if valor_csrf > 0 or (tipo_retencao_csrf and str(tipo_retencao_csrf).strip()):
                    if abs(valor_csrf - valor_esperado) > 0.01:
                        alertas.append(
                            f"CSRF retido divergente para código {codigo_norm}. Esperado: {valor_esperado:.2f} ({exp_csrf_percent}%). Encontrado: {valor_csrf:.2f}"
                        )
                else:
                    if valor_esperado > 0.01:
                        alertas.append(
                            f"CSRF devido e não retido para código {codigo_norm}. Deveria ser: {valor_esperado:.2f} ({exp_csrf_percent}%)"
                        )
        elif exp_csrf_flag == 'DEPENDE' and (valor_csrf > 0 or (tipo_retencao_csrf and str(tipo_retencao_csrf).strip())):
            alertas.append(
                f"CSRF na planilha = DEPENDE para código {codigo_norm}. Revisar regra/observação."
            )

        # INSS (planilha só diz SIM/NÃO/DEPENDE; não calcula %)
        if exp_inss_bool is True and valor_inss <= 0:
            alertas.append(
                f"INSS esperado (planilha=SIM) para código {codigo_norm}, mas veio 0.00."
            )
        elif exp_inss_bool is False and valor_inss > 0:
            alertas.append(
                f"INSS não esperado (planilha=NÃO) para código {codigo_norm}, mas veio {valor_inss:.2f}."
            )
        elif exp_inss_flag == 'DEPENDE' and valor_inss > 0:
            alertas.append(
                f"INSS na planilha = DEPENDE para código {codigo_norm}. Revisar regra/observação."
            )

        return {'alertas': alertas, 'correcoes': correcoes}

    
    def extract_value(self, element, path, default=''):
        try:
            namespaced_path = './/ns:' + path.replace('/', '/ns:')
            result = element.find(namespaced_path, self.namespaces)
            if result is not None and result.text:
                return result.text.strip()
            result = element.find(f'.//{path}')
            if result is not None and result.text:
                return result.text.strip()
            return default
        except:
            return default
    
    def extract_tributacao_info(self, element):
        trib_info = {
            'vISSQN': '0.00',
            'tpRetISSQN': '1',
            'vRetIRRF': '0.00',
            'pRetIRRF': '0.00',
            'vRetCSLL': '0.00',
            'vRetINSS': '0.00',
            'vBC': '0.00',
            'vCP': '0.00',
            'vRetCP': '0.00'
        }
        
        # paths for extraction
        paths = [
            ('trib/tribMun/vISSQN', 'vISSQN'),
            ('valores/vISSQN', 'vISSQN'),
            ('vISSQN', 'vISSQN'),
            ('trib/tribMun/tpRetISSQN', 'tpRetISSQN'),
            ('tribFed/vRetIRRF', 'vRetIRRF'),
            ('tribFed/pRetIRRF', 'pRetIRRF'),
            ('tribFed/vRetCSLL', 'vRetCSLL'),
            ('tribFed/vRetINSS', 'vRetINSS'),
            ('tribFed/cp/vCP', 'vCP'),
            ('cp/vCP', 'vCP'),
            ('tribFed/vRetCP', 'vRetCP'),
            ('valores/vBC', 'vBC'),
            ('vBC', 'vBC'),
        ]
        
        for xpath, key in paths:
            value = self.extract_value(element, xpath)
            if value and value != '0.00':
                trib_info[key] = value
        
        return trib_info
    
    def parse_xml(self, xml_content):
        """Analisa um arquivo XML e extrai os dados"""
        try:
            root = ET.fromstring(xml_content)
            
            inf_nfse = root.find('.//ns:infNFSe', self.namespaces)
            if inf_nfse is None:
                inf_nfse = root.find('.//infNFSe')
                if inf_nfse is None:
                    inf_nfse = root.find('.//ns:infDPS', self.namespaces)
                    if inf_nfse is None:
                        inf_nfse = root.find('.//infDPS')
                        if inf_nfse is None:
                            return None
            
            competencia = self.extract_value(inf_nfse, 'dCompet')

            # Datas do XML padrão nacional:
            # - dhProc: data/hora de emissão/processamento da NFS-e (data correta para mês/período)
            # - dhEmi: data/hora de emissão da DPS (fallback)
            # - dCompet: competência (último fallback)
            dh_proc = self.extract_value(inf_nfse, 'dhProc') or self.extract_value(root, 'dhProc')

            # Data de Emissão (PLANILHA) deve vir da emissão real (DPS): dhEmi (ou dEmi como fallback).
            # NUNCA usar dhProc para preencher "Data de Emissão" na planilha.
            dh_emi = (
                self.extract_value(inf_nfse, 'dhEmi')
                or self.extract_value(root, 'dhEmi')
                or self.extract_value(inf_nfse, 'dEmi')
                or self.extract_value(root, 'dEmi')
            )

            def _date_part(v: str) -> str:
                if not v:
                    return ''
                v = v.strip()
                # ISO 8601 -> YYYY-MM-DD
                return v[:10] if len(v) >= 10 else (v.split('T')[0] if 'T' in v else v)

            # Planilha: dhEmi (emissão). Fallback: dCompet (se existir).
            data_nota = _date_part(dh_emi) or competencia

            # Mantemos a coluna "Data de Emissão" como a data de referência da nota (prioriza dhProc)
            data_emissao = data_nota

            # Chave de acesso (quando existir no XML)
            # A chave pode estar no atributo 'Id' do elemento infNFSe (ex: NFS21113002250810235000117000000000010426032336425030)
            # ou em elementos como chNFSe, chNfse, chNFS, chNFS-e
            chave_acesso = ''
            
            # Primeiro tenta extrair do atributo 'Id' do elemento infNFSe
            if inf_nfse is not None:
                chave_acesso = inf_nfse.get('Id', '')
                if chave_acesso:
                    # Remove prefixo comum como "NFS" se presente
                    if chave_acesso.startswith('NFS'):
                        chave_acesso = chave_acesso[3:]
                    # Remove caracteres não numéricos
                    chave_acesso = re.sub(r'[^\d]', '', chave_acesso)
                    print(f"  🔑 Chave de Acesso extraída do atributo Id: {chave_acesso}")
            
            # Se não encontrou no atributo, tenta elementos de texto
            if not chave_acesso:
                chave_acesso = (
                    self.extract_value(inf_nfse, 'chNFSe') or self.extract_value(root, 'chNFSe') or
                    self.extract_value(inf_nfse, 'chNfse') or self.extract_value(root, 'chNfse') or
                    self.extract_value(inf_nfse, 'chNFS') or self.extract_value(root, 'chNFS') or
                    self.extract_value(inf_nfse, 'chNFS-e') or self.extract_value(root, 'chNFS-e')
                )
                if chave_acesso:
                    print(f"  🔑 Chave de Acesso extraída do elemento: {chave_acesso}")
            
            if not chave_acesso:
                print(f"  ⚠️ ATENÇÃO: Chave de Acesso NÃO encontrada no XML!")

            municipio = self.extract_value(inf_nfse, 'xLocEmi')
            
            # ================================================================
            # Extrair CNPJ/CPF baseado no tipo de nota:
            # - tomados (Recebidas): usa emitente (prestador de serviço)
            # - prestados (Emitidas): usa tomador (cliente que recebeu o serviço)
            # ================================================================
            
            if self.tipo_nota == 'prestados':
                # Para notas emitidas (prestados), extrai do TOMADOR
                # Tenta várias tags possíveis: tomador/CNPJ, toma/CNPJ, destinatario/CNPJ
                cnpj_emitente = (
                    self.extract_value(inf_nfse, 'tomador/CNPJ') or
                    self.extract_value(inf_nfse, 'toma/CNPJ') or
                    self.extract_value(inf_nfse, 'destinatario/CNPJ')
                )
                cpf_emitente = None
                
                if not cnpj_emitente:
                    # Tentar CPF em vez de CNPJ
                    cpf_emitente = (
                        self.extract_value(inf_nfse, 'tomador/CPF') or
                        self.extract_value(inf_nfse, 'toma/CPF') or
                        self.extract_value(inf_nfse, 'destinatario/CPF')
                    )
                    if cpf_emitente:
                        cnpj_emitente = cpf_emitente
                
                # Extrair razão social do tomador para notas prestadas
                razao_social = (
                    self.extract_value(inf_nfse, 'tomador/xNome') or
                    self.extract_value(inf_nfse, 'toma/xNome') or
                    self.extract_value(inf_nfse, 'destinatario/xNome')
                )
            else:
                # Para notas tomadas (recebidas), extrai do EMITENTE (lógica original)
                # Extrair CNPJ ou CPF do emitente (ordem EXATA: emit/CNPJ -> emit/CPF -> prest/CNPJ -> prest/CPF -> toma/CNPJ -> toma/CPF)
                cnpj_emitente = (
                    self.extract_value(inf_nfse, 'emit/CNPJ') or
                    self.extract_value(inf_nfse, 'prest/CNPJ') or
                    self.extract_value(inf_nfse, 'toma/CNPJ')
                )
                cpf_emitente = None
                
                if not cnpj_emitente:
                    # Tentar CPF em vez de CNPJ
                    cpf_emitente = (
                        self.extract_value(inf_nfse, 'emit/CPF') or
                        self.extract_value(inf_nfse, 'prest/CPF') or
                        self.extract_value(inf_nfse, 'toma/CPF')
                    )
                    if cpf_emitente:
                        cnpj_emitente = cpf_emitente
                
                # Para notas tomadas, usa emit/xNome (lógica original)
                razao_social = self.extract_value(inf_nfse, 'emit/xNome')
                if not razao_social:
                    razao_social = self.extract_value(inf_nfse, 'prest/xNome')
                if not razao_social:
                    razao_social = self.extract_value(inf_nfse, 'toma/xNome')
            
            # Criar doc_emitente_raw (somente dígitos, removendo ., / e -)
            doc_emitente_raw = ''
            if cnpj_emitente:
                doc_emitente_raw = re.sub(r'[^\d]', '', str(cnpj_emitente))
            
            # Verificar se é CPF-only (11 dígitos)
            is_cpf_only = len(doc_emitente_raw) == 11
            
            # Se doc_emitente_raw ficou vazio, logar erro (será verificado depois com nome do arquivo)
            # Armazenar para possível logging posterior
            has_empty_doc = not doc_emitente_raw
            
            n_nfse = self.extract_value(inf_nfse, 'nNFSe')
            if not n_nfse:
                n_nfse = self.extract_value(inf_nfse, 'nDPS')
            
            valor_total = self.extract_value(inf_nfse, 'valores/vServ')
            if not valor_total:
                valor_total = self.extract_value(inf_nfse, 'vServPrest/vServ')
            if not valor_total:
                valor_total = self.extract_value(inf_nfse, 'valores/vBC')
            if not valor_total:
                valor_total = self.extract_value(inf_nfse, 'vServ')
            
            valor_liquido = self.extract_value(inf_nfse, 'valores/vLiq', '0.00')
            if valor_liquido == '0.00':
                valor_liquido = self.extract_value(inf_nfse, 'vLiq', '0.00')
            
            trib_info = self.extract_tributacao_info(inf_nfse)
            
            tp_ret_issqn = trib_info['tpRetISSQN']
            iss_valor = trib_info['vISSQN']
            iss_retencao = iss_valor if tp_ret_issqn == '2' else '0.00'
            
            # CSRF: vem diretamente do campo vRetCSLL do XML
            csrf_valor = trib_info.get('vRetCSLL', '0.00')
            
            irrf_valor = trib_info['vRetIRRF']
            percentual_irrf = trib_info['pRetIRRF']
            
            if float(irrf_valor) > 0 and float(percentual_irrf) == 0:
                vbc_valor = trib_info['vBC']
                if float(vbc_valor) > 0:
                    percentual_calculado = (float(irrf_valor) / float(vbc_valor)) * 100
                    percentual_irrf = f"{percentual_calculado:.2f}"
                    print(f"  📊 Percentual IRRF calculado: {percentual_irrf}% (IRRF: {irrf_valor} / Base: {vbc_valor})")
            
            if float(trib_info['vRetCP']) > 0:
                inss_valor = trib_info['vRetCP']
                print(f"  📊 INSS extraído do campo vRetCP (XML 2026): {inss_valor}")
            elif float(trib_info['vCP']) > 0:
                inss_valor = trib_info['vCP']
                print(f"  📊 INSS extraído do campo vCP: {inss_valor}")
            else:
                inss_valor = trib_info['vRetINSS']
                print(f"  📊 INSS extraído do campo vRetINSS: {inss_valor}")
            
            vbc_valor = trib_info['vBC']
            
            incidencia_iss = self.extract_value(inf_nfse, 'xLocIncid')
            if not incidencia_iss:
                incidencia_iss = self.extract_value(inf_nfse, 'xLocPrestacao')
            if not incidencia_iss:
                incidencia_iss = municipio
            
            # ---------- CÓDIGO DE SERVIÇO COM NORMALIZAÇÃO AVANÇADA ----------
            cod_servico_raw = self.extract_value(inf_nfse, 'cServ/cTribNac')
            if not cod_servico_raw:
                cod_servico_raw = self.extract_value(inf_nfse, 'cTribNac')
            
            cod_servico_formatado = self._extrair_codigo_lc116(cod_servico_raw)
            cod_servico_normalizado = self.normalizar_codigo_servico(cod_servico_formatado)
            print(f"  🔄 Código normalizado: {cod_servico_raw} -> {cod_servico_formatado} -> {cod_servico_normalizado}")
            # --------------------------------------------------------
            
            descricao_servico = self.extract_value(inf_nfse, 'xDescServ')
            codigo_nbs = self.extract_value(inf_nfse, 'cNBS')
            codigo_cnae = self.extract_value(inf_nfse, 'cCnae')
            descricao_cnae = self.extract_value(inf_nfse, 'xCnae')
            
            # Mapeamento do Simples Nacional
            simples_nacional = self.extract_value(inf_nfse, 'prest/regTrib/opSimpNac')
            if simples_nacional == '1':
                regime = 'Não optante'
            elif simples_nacional == '2':
                regime = 'MEI'
            elif simples_nacional == '3':
                regime = 'Optante S.N'
            else:
                reg_trib = self.extract_value(inf_nfse, 'prest/regTrib/regApTribSN')
                if reg_trib == '1':
                    regime = 'Simples Nacional'
                else:
                    regime = 'Não optante'
            
            data_pagamento = data_emissao
            consulta_api = ''
            
            data = {
                'Competência': competencia,
                'Município': municipio,
                'Chave de Acesso': chave_acesso,
                'Data/Hora Emissão NFS-e': dh_proc,
                'Data de Emissão': data_emissao,
                'CNPJ/CPF': self.format_cnpj(doc_emitente_raw),
                'Razão Social': razao_social,
                'N° Documento': n_nfse,
                'Valor Total': self.format_currency(self._to_float(valor_total)),
                'Valor B/C': self.format_currency(self._to_float(vbc_valor)),
                'CSRF': self.format_currency(self._to_float(csrf_valor)),
                'IRRF': self.format_currency(self._to_float(irrf_valor)),
                'Percentual IRRF': percentual_irrf if percentual_irrf else '0.00',
                'INSS': self.format_currency(self._to_float(inss_valor)),
                'ISS': self.format_currency(self._to_float(iss_retencao)),
                'Valor Líquido': self.format_currency(self._to_float(valor_liquido)),
                'Incidência do ISS': incidencia_iss,
                'Data do pagamento': data_pagamento,
                'Código de serviço': cod_servico_normalizado,
                'Descrição do Serviço': descricao_servico,
                'Código NBS': codigo_nbs,
                'Código CNAE': codigo_cnae,
                'Descrição CNAE': descricao_cnae,
                'Simples Nacional / XML': regime,
                'Consulta Simples API': consulta_api,
                '_CNPJ_Raw': doc_emitente_raw,
                '_Is_CPF_Only': is_cpf_only,
                '_Has_Empty_Doc': has_empty_doc,
                '_Arquivo_Origem': ''  # será preenchido em process_multiple_files
            }
            
            # 🔎 AUDITORIA (MODO ALERTA): NÃO SOBRESCREVE VALORES DO XML
            print(f"\n📋 Processando código de serviço: {cod_servico_normalizado}")
            print(f"📋 Regime (XML): {regime}")
            # 2️⃣ GERA ALERTAS (comparação XML x regra) - sem alterar valores
            correcoes = self.verificar_correcoes(data)
            data['_Correcoes'] = correcoes

            # 3️⃣ STATUS DE AUDITORIA + ALERTAS FISCAIS (para consumo na planilha)
            status_sn, status_csrf, status_irrf, status_inss, alertas_fiscais = self.determinar_status_auditoria(data)
            data['Status Simples Nacional'] = status_sn
            data['Status CSRF'] = status_csrf
            data['Status IRRF'] = status_irrf
            data['Status INSS'] = status_inss
            data['Alertas Fiscais'] = alertas_fiscais

            return data
            
        except Exception as e:
            print(f"Erro ao analisar XML: {str(e)}")
            traceback.print_exc()
            return None
    
    # ------------------------------------------------------------------
    # determinar_status_auditoria e demais métodos auxiliares (mantidos)
    # ------------------------------------------------------------------
    def _categoria_simples(self, texto: str) -> str:
        """Normaliza o texto de regime (XML/API) para categorias comparáveis."""
        if not texto:
            return ""
        t = str(texto).strip()
        # remove acentos para facilitar matching
        t_noacc = ''.join(ch for ch in unicodedata.normalize('NFKD', t) if not unicodedata.combining(ch))
        u = t_noacc.upper()
        if "MEI" in u:
            return "MEI"
        # API/strings comuns
        if "NAO" in u or "NÃO" in t.upper():
            if "OPTANTE" in u or "SIMPLES" in u:
                return "NAO_OPTANTE"
        if "OPTANTE" in u or "SIMPLES" in u:
            return "OPTANTE"
        return u  # fallback

    def _status_simples_nacional(self, regime_xml: str, regime_api: str) -> str:
        """Retorna Correto/Divergente/Pendente conforme regra acordada (MEI + Optante = correto)."""
        if not (regime_api or '').strip():
            return "Pendente"
        cat_xml = self._categoria_simples(regime_xml)
        cat_api = self._categoria_simples(regime_api)
        # Regra especial: XML=MEI e API=OPTANTE => correto
        if cat_xml == "MEI" and cat_api == "OPTANTE":
            return "Correto"
        # Comparação por categoria (quando dá)
        if cat_xml and cat_api and cat_xml == cat_api:
            return "Correto"
        # Fallback: comparação de string (case-insensitive, sem acento)
        def _norm(s: str) -> str:
            s = '' if s is None else str(s).strip()
            s = ''.join(ch for ch in unicodedata.normalize('NFKD', s) if not unicodedata.combining(ch))
            return s.upper()
        return "Correto" if _norm(regime_xml) == _norm(regime_api) else "Divergente"

    def determinar_status_auditoria(self, dados):
        """
        Define os status finais para a planilha (sem corrigir valores do XML).

        - Status Simples Nacional:
            * Pendente: quando não existe "Consulta Simples API"
            * Correto/Divergente: quando existe e difere do "Simples Nacional / XML"
        - Status CSRF / IRRF:
            * Divergente quando "Alertas Fiscais" contiver indícios relacionados ao imposto
        - Alertas Fiscais:
            * Usa o resultado gerado por verificar_correcoes (modo alerta)
            * Se houver divergência de Simples via API, adiciona mensagem complementar.
        
        REGRAS ESPECIAIS:
        - MEI (XML + API): Não deve ter nenhuma retenção de imposto
        - Optante do Simples (XML + API): IRRF/CSRF não devem ter retenção, mas ISS segue regras
        """
        correcoes = (dados.get('_Correcoes', '') or '').strip()
        codigo_servico = (dados.get('Código de serviço', '') or '').strip()

        # Status SN (pode ser sobrescrito depois pela classe ComAPI)
        regime_xml = (dados.get('Simples Nacional / XML', '') or '').strip()
        regime_api = (dados.get('Consulta Simples API', '') or '').strip()

        status_sn = self._status_simples_nacional(regime_xml, regime_api)

        # Status IRRF / CSRF / INSS
        status_irrf = "Divergente" if "IRRF" in correcoes else "Correto"
        status_csrf = "Divergente" if "CSRF" in correcoes else "Correto"
        status_inss = "Divergente" if "INSS" in correcoes else "Correto"

        alertas_fiscais = correcoes

        # ================================================================
        # REGRA ESPECIAL PARA MEI/MEI
        # Se MEI (XML) E MEI (API), não deve ter nenhuma retenção de imposto
        # ================================================================
        cat_xml = self._categoria_simples(regime_xml)
        cat_api = self._categoria_simples(regime_api)
        
        if cat_xml == "MEI" and cat_api == "MEI":
            # Verifica valores dos impostos
            valor_csrf = self._to_float(dados.get('CSRF', 0))
            valor_irrf = self._to_float(dados.get('IRRF', 0))
            valor_inss = self._to_float(dados.get('INSS', 0))
            valor_iss = self._to_float(dados.get('ISS', 0))
            
            # Caso 1: todos os impostos = 0 → CORRETO
            if valor_csrf == 0 and valor_irrf == 0 and valor_inss == 0 and valor_iss == 0:
                # Todos os status como Correto
                status_csrf = "Correto"
                status_irrf = "Correto"
                status_inss = "Correto"
                
                # Adiciona alerta informativo (sem remover outros alertas existentes)
                alerta_mei = "MEI Correto"
                if alertas_fiscais:
                    alertas_fiscais = alertas_fiscais + " | " + alerta_mei
                else:
                    alertas_fiscais = alerta_mei
            else:
                # Caso 2: algum imposto > 0 → DIVERGENTE
                status_csrf = "Divergente"
                status_irrf = "Divergente"
                status_inss = "Divergente"
                
                # Identifica quais impostos têm valores
                impostos_retidos = []
                if valor_csrf > 0:
                    impostos_retidos.append(f"CSRF ({valor_csrf:.2f})")
                if valor_irrf > 0:
                    impostos_retidos.append(f"IRRF ({valor_irrf:.2f})")
                if valor_inss > 0:
                    impostos_retidos.append(f"INSS ({valor_inss:.2f})")
                if valor_iss > 0:
                    impostos_retidos.append(f"ISS ({valor_iss:.2f})")
                
                # Adiciona alerta de divergência específico
                alerta_mei = "MEI DIVERGENTE - IMPOSTO RETIDO (" + ", ".join(impostos_retidos) + ")"
                if alertas_fiscais:
                    alertas_fiscais = alertas_fiscais + " | " + alerta_mei
                else:
                    alertas_fiscais = alerta_mei

        # ================================================================
        # REGRA ESPECIAL PARA OPTANTE DO SIMPLES/OPTANTE DO SIMPLES
        # Se Optante (XML) E Optante (API), não deve ter IRRF/CSRF retido
        # (INSS pode reter conforme Anexo IV)
        # ================================================================
        if cat_xml == "OPTANTE" and cat_api == "OPTANTE":
            # Verifica valores de IRRF e CSRF (não deve ter)
            valor_csrf = self._to_float(dados.get('CSRF', 0))
            valor_irrf = self._to_float(dados.get('IRRF', 0))
            
            # Caso 1: IRRF e CSRF = 0 → CORRETO para esses impostos
            if valor_csrf == 0 and valor_irrf == 0:
                status_csrf = "Correto"
                status_irrf = "Correto"
                
                # Adiciona alerta informativo (sem remover outros alertas existentes)
                alerta_optante = "Optante Simples Correto"
                if alertas_fiscais:
                    alertas_fiscais = alertas_fiscais + " | " + alerta_optante
                else:
                    alertas_fiscais = alerta_optante
            else:
                # Caso 2: IRRF ou CSRF > 0 → DIVERGENTE para esses impostos
                if valor_irrf > 0:
                    status_irrf = "Divergente"
                if valor_csrf > 0:
                    status_csrf = "Divergente"
                
                # Identifica quais impostos têm valores
                impostos_retidos = []
                if valor_csrf > 0:
                    impostos_retidos.append(f"CSRF ({valor_csrf:.2f})")
                if valor_irrf > 0:
                    impostos_retidos.append(f"IRRF ({valor_irrf:.2f})")
                
                # Adiciona alerta de divergência específico
                alerta_optante = "Optante Simples DIVERGENTE - IMPOSTO RETIDO (" + ", ".join(impostos_retidos) + ")"
                if alertas_fiscais:
                    alertas_fiscais = alertas_fiscais + " | " + alerta_optante
                else:
                    alertas_fiscais = alerta_optante

        if status_sn == "Divergente":
            sn_msg = f"Simples Nacional divergente: XML={regime_xml} vs API={regime_api}"
            alertas_fiscais = (alertas_fiscais + " | " + sn_msg) if alertas_fiscais else sn_msg

        return status_sn, status_csrf, status_irrf, status_inss, alertas_fiscais


    def verificar_codigos_sem_regras(self, dados_processados):
        codigos_encontrados = set()
        codigos_sem_regra = set()
        
        for item in dados_processados:
            codigo = item.get('Código de serviço', '')
            if codigo:
                codigos_encontrados.add(codigo)
                if (codigo not in self.REGRAS_RETENCOES):
                    codigos_sem_regra.add(codigo)
        
        if codigos_sem_regra:
            print(f"\n{'='*60}")
            print("⚠️  CÓDIGOS SEM REGRAS DEFINIDAS (mantidos do XML):")
            print("="*60)
            for codigo in sorted(codigos_sem_regra):
                print(f"  - {codigo}")
            print("="*60)
        
        return list(codigos_sem_regra)
    
    def process_multiple_files(self, file_paths):
        all_data = []
        errors = []
        
        for file_path in file_paths:
            try:
                with open(file_path, 'r', encoding='utf-8') as file:
                    xml_content = file.read()
                
                data = self.parse_xml(xml_content)
                if data:
                    data['_Arquivo_Origem'] = os.path.basename(file_path)
                    data['_arquivo_origem'] = os.path.basename(file_path)
                    if data['CNPJ/CPF'] and data['N° Documento']:
                        all_data.append(data)
                        print(f"✓ Processado: {os.path.basename(file_path)}")
                    else:
                        errors.append(f"Campos obrigatórios faltando: {os.path.basename(file_path)}")
                else:
                    errors.append(f"Erro no processamento: {os.path.basename(file_path)}")
                    
            except Exception as e:
                error_msg = f"Erro ao ler arquivo {os.path.basename(file_path)}: {str(e)}"
                errors.append(error_msg)
                print(f"✗ {error_msg}")
        
        if errors:
            print(f"\n⚠️  {len(errors)} erro(s) encontrado(s):")
            for error in errors[:5]:
                print(f"  - {error}")
            if len(errors) > 5:
                print(f"  ... e mais {len(errors) - 5} erro(s)")
        
        self.verificar_codigos_sem_regras(all_data)
        return all_data
    
    def save_to_excel(self, data, output_path):
        """Salva dados em arquivo Excel com auditoria fiscal"""
        if not data:
            print("Nenhum dado para salvar.")
            return False
        
        df = pd.DataFrame(data)
        
        # ================================================================
        # FUNÇÃO AUXILIAR: VERIFICAR CAMPOS AUSENTES NO XML
        # ================================================================
        def verificar_campos_ausentes(row):
            """
            Verifica campos importantes que não vieram preenchidos no XML.
            Retorna string com nomes dos campos ausentes separados por '; '
            ou 'Nenhum' se todos estiverem presentes.
            """
            campos_ausentes = []
            
            # Mapeamento de campos a verificar (nome na planilha: é_monetario)
            campos_verificar = [
                ('Valor B/C', True),
                ('CSRF', True),
                ('IRRF', True),
                ('INSS', True),
                ('ISS', True),
                ('Valor Total', True),
                ('Valor Líquido', True),
                ('CNPJ/CPF', False),
                ('Percentual IRRF', False),
                ('Incidência do ISS', False),
                ('Código de serviço', False),
                ('Código CNAE', False),
                ('Código NBS', False),
            ]
            
            for campo_nome, e_monenario in campos_verificar:
                # Regra 1: Coluna não existe no dataframe
                if campo_nome not in row.index:
                    campos_ausentes.append(campo_nome)
                    continue
                
                valor = row[campo_nome]
                
                # Regra 2: Valor é NaN
                if pd.isna(valor):
                    campos_ausentes.append(campo_nome)
                    continue
                
                # Regra 3: String vazia ou apenas espaços
                if isinstance(valor, str):
                    if valor.strip() == '':
                        campos_ausentes.append(campo_nome)
                        continue
                
                # Regra 4: Valor numérico igual a 0 (apenas para campos monetários)
                if e_monenario:
                    try:
                        valor_numerico = float(valor)
                        if valor_numerico == 0:
                            campos_ausentes.append(campo_nome)
                    except (ValueError, TypeError):
                        # Se não conseguir converter para número, não considera ausente
                        pass
            
            if campos_ausentes:
                return "; ".join(campos_ausentes)
            else:
                return "Nenhum"
        
        # Aplicar a verificação de campos ausentes
        df["Campos ausentes no XML"] = df.apply(verificar_campos_ausentes, axis=1)
        
        # ================================================================
        # COLUNAS: VALOR LÍQUIDO CORRETO E STATUS VALOR LÍQUIDO
        # ================================================================
        # Converter colunas para numeric antes do cálculo
        df['Valor Total'] = pd.to_numeric(df['Valor Total'], errors='coerce').fillna(0.0)
        df['CSRF'] = pd.to_numeric(df['CSRF'], errors='coerce').fillna(0.0)
        df['IRRF'] = pd.to_numeric(df['IRRF'], errors='coerce').fillna(0.0)
        df['INSS'] = pd.to_numeric(df['INSS'], errors='coerce').fillna(0.0)
        df['ISS'] = pd.to_numeric(df['ISS'], errors='coerce').fillna(0.0)
        df['Valor Líquido'] = pd.to_numeric(df['Valor Líquido'], errors='coerce').fillna(0.0)
        
        # ================================================================
        # COLUNA: STATUS BASE DE CÁLCULO
        # Se Valor B/C > 0 = Correto
        # Se Valor B/C = 0 E é MEI = Correto (MEI com base zerada está correto)
        # Se Valor B/C = 0 E não é MEI = Divergente
        # ================================================================
        df['Valor B/C'] = pd.to_numeric(df['Valor B/C'], errors='coerce').fillna(0.0)
        
        def _status_base_calculo(row):
            status = compute_base_calculation_status(row['Valor B/C'], row['Valor Total'])
            return 'Correto' if status == 'ok' else 'Divergente'
        
        df['Status Base de Cálculo'] = df.apply(_status_base_calculo, axis=1)
        # Restaurar formato original de string para a coluna
        df['Valor B/C'] = df['Valor B/C'].apply(self.format_currency)
        
        # Calcular Valor Líquido Correto = Valor Total - (CSRF + IRRF + INSS + ISS)
        df['Valor Líquido Correto'] = df['Valor Total'] - (df['CSRF'] + df['IRRF'] + df['INSS'] + df['ISS'])
        
        # Comparar com tolerância de 0.01
        df['Status Valor Líquido'] = df.apply(
            lambda row: 'Correto' if abs(row['Valor Líquido'] - row['Valor Líquido Correto']) <= 0.01 else 'Divergente',
            axis=1
        )
        
        # Restaurar formato original de string para as colunas
        df['Valor Total'] = df['Valor Total'].apply(self.format_currency)
        df['CSRF'] = df['CSRF'].apply(self.format_currency)
        df['IRRF'] = df['IRRF'].apply(self.format_currency)
        df['INSS'] = df['INSS'].apply(self.format_currency)
        df['ISS'] = df['ISS'].apply(self.format_currency)
        df['Valor Líquido'] = df['Valor Líquido'].apply(self.format_currency)
        df['Valor Líquido Correto'] = df['Valor Líquido Correto'].apply(self.format_currency)
        # ================================================================
        
        debug_cols = ['_CNPJ_Raw', '_Alertas', '_Correcoes_Sugeridas', '_Correcoes']
        for col in debug_cols:
            if col in df.columns:
                df = df.drop(columns=[col])
        
        df['_Divergente'] = df.apply(
            lambda row: compute_final_note_status(
                {
                    'status_simples_nacional': row.get('Status Simples Nacional'),
                    'status_csrf': row.get('Status CSRF'),
                    'status_irrf': row.get('Status IRRF'),
                    'status_inss': row.get('Status INSS'),
                    'status_valor_liquido': row.get('Status Valor Líquido'),
                }
            ) == 'divergente',
            axis=1
        )
        
        df_divergentes = df[df['_Divergente']].copy()
        df_corretas = df[~df['_Divergente']].copy()
        
        df = df.drop(columns=['_Divergente'])
        df_divergentes = df_divergentes.drop(columns=['_Divergente'])
        df_corretas = df_corretas.drop(columns=['_Divergente'])
        
        column_order = [
            'Competência', 'Município', 'Chave de Acesso', 'Data de Emissão', 'CNPJ/CPF', 
            'Razão Social', 'N° Documento', 'Valor Total', 'Valor B/C', 'Status Base de Cálculo',
            'CSRF', 'IRRF', 'Percentual IRRF', 'INSS', 'ISS', 
            'Valor Líquido', 'Valor Líquido Correto', 'Status Valor Líquido', 'Campos ausentes no XML',
            'Incidência do ISS', 'Data do pagamento', 
            'Código de serviço', 'Descrição do Serviço', 'Código NBS',
            'Código CNAE', 'Descrição CNAE',
            'Simples Nacional / XML', 'Consulta Simples API',
            'Status Simples Nacional', 'Status CSRF', 'Status IRRF', 'Status INSS',
            'Alertas Fiscais'
        ]
        
        existing_columns = [col for col in column_order if col in df.columns]
        
        df = df[existing_columns]
        if not df_divergentes.empty:
            df_divergentes = df_divergentes[existing_columns]
        if not df_corretas.empty:
            df_corretas = df_corretas[existing_columns]
        
        if 'Data de Emissão' in df.columns:
            for dataframe in [df, df_divergentes, df_corretas]:
                if not dataframe.empty:
                    dataframe['Data de Emissão'] = pd.to_datetime(dataframe['Data de Emissão'].astype(str).str[:10], errors='coerce', format='%Y-%m-%d')
                    dataframe.sort_values('Data de Emissão', inplace=True)
                    dataframe['Data de Emissão'] = dataframe['Data de Emissão'].dt.strftime('%Y-%m-%d')
        
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Todas as Notas', index=False)
            
            if not df_divergentes.empty:
                df_divergentes.to_excel(writer, sheet_name='Notas Divergentes', index=False)
            else:
                pd.DataFrame({'Mensagem': ['Nenhuma nota divergente encontrada']}).to_excel(
                    writer, sheet_name='Notas Divergentes', index=False)
            
            if not df_corretas.empty:
                df_corretas.to_excel(writer, sheet_name='Notas Corretas', index=False)
            else:
                pd.DataFrame({'Mensagem': ['Nenhuma nota correta encontrada']}).to_excel(
                    writer, sheet_name='Notas Corretas', index=False)
            
            workbook = writer.book
            currency_format = workbook.add_format({'num_format': '#,##0.00'})
            date_format = workbook.add_format({'num_format': 'yyyy-mm-dd'})
            text_format = workbook.add_format()
            descricao_format = workbook.add_format({'text_wrap': True})
            percent_format = workbook.add_format({'num_format': '0.00%'})
            status_correto_format = workbook.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'})
            status_divergente_format = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
            
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                if sheet_name == 'Todas as Notas':
                    current_df = df
                elif sheet_name == 'Notas Divergentes':
                    current_df = df_divergentes if not df_divergentes.empty else pd.DataFrame()
                elif sheet_name == 'Notas Corretas':
                    current_df = df_corretas if not df_corretas.empty else pd.DataFrame()
                else:
                    # Não há mais aba/df de regras no output
                    current_df = pd.DataFrame()
                
                if current_df.empty:
                    worksheet.set_column('A:A', 50)
                    continue
                
                for i, col in enumerate(current_df.columns):
                    max_length = max(
                        current_df[col].astype(str).map(len).max(),
                        len(col)
                    ) + 2
                    
                    if col in ['Descrição do Serviço', 'Descrição CNAE']:
                        column_width = min(max_length, 80)
                    elif col in ['Alertas Fiscais']:
                        column_width = min(max_length, 50)
                    else:
                        column_width = min(max_length, 30)
                    
                    if any(keyword in col for keyword in ['Valor', 'CSRF', 'IRRF', 'INSS', 'ISS', 'Valor B/C']):
                        worksheet.set_column(i, i, column_width, currency_format)
                    elif 'Data' in col or 'Competência' in col:
                        worksheet.set_column(i, i, column_width, date_format)
                    elif col == 'Percentual IRRF':
                        col_data = current_df[col]
                        try:
                            numeric_vals = col_data.astype(float) / 100
                        except:
                            numeric_vals = col_data
                        worksheet.set_column(i, i, column_width, percent_format)
                        for row_num, val in enumerate(numeric_vals):
                            worksheet.write_number(row_num + 1, i, val, percent_format)
                    elif col in ['Descrição do Serviço', 'Descrição CNAE']:
                        worksheet.set_column(i, i, column_width, descricao_format)
                    elif 'Status' in col:
                        for row_num in range(1, len(current_df) + 1):
                            cell_value = current_df.iloc[row_num-1][col] if row_num-1 < len(current_df) else ''
                            if cell_value == 'Correto':
                                worksheet.write(row_num, i, cell_value, status_correto_format)
                            elif cell_value in ['Divergente', 'Divergência']:
                                worksheet.write(row_num, i, cell_value, status_divergente_format)
                            else:
                                worksheet.write(row_num, i, cell_value, text_format)
                        worksheet.set_column(i, i, column_width)
                    else:
                        worksheet.set_column(i, i, column_width, text_format)
                
                worksheet.autofilter(0, 0, len(current_df), len(current_df.columns) - 1)
                worksheet.freeze_panes(1, 0)
        
        return True


class NFSeXMLConverterComAPI(NFSeXMLConverter):
    def __init__(self, tipo_nota: str = "tomados", consultar_api: bool = True):
        """
        Inicializa o conversor com suporte a consulta de API.
        
        Args:
            tipo_nota: 'tomados' (Recebidas) ou 'prestados' (Emitidas)
            consultar_api: se True, consulta CNPJ via API Invertexto
        """
        super().__init__(tipo_nota=tipo_nota)
        self.consultar_api = consultar_api
        if consultar_api:
            # Invertexto (consulta CNPJ). Use delay_seconds para controlar o ritmo.
            self.cnpj_consultor = CNPJConsultor(delay_seconds=1.0)
    
    def process_multiple_files(self, file_paths):
        """Compat: main.py chama process_multiple_files."""
        return super().process_multiple_files(file_paths)
    
    def consultar_cnpjs_em_lote(self, dados):
        """Consulta CNPJs em lote para preencher campo de Simples Nacional e CNAE via API"""
        if not self.consultar_api:
            print("Consulta de API desabilitada.")
            return dados
        
        print(f"\n{'='*60}")
        print("INICIANDO CONSULTA DE SIMPLES NACIONAL E CNAE VIA API")
        print("="*60)
        print("ℹ️  Fonte: Invertexto (consulta CNPJ).")
        print("⏳ Throttle: 1 consulta por CNPJ e espera configurável quando não vier do cache")
        print("="*60)
        
        total_cnpjs = 0
        consultas_realizadas = 0
        consultas_cache = 0
        
        cnpjs_unicos = {}
        for i, item in enumerate(dados):
            cnpj_raw = item.get('_CNPJ_Raw', '')
            cnpj_formatado = item['CNPJ/CPF']
            
            if cnpj_raw:
                cnpj_para_consulta = cnpj_raw
            else:
                cnpj_para_consulta = re.sub(r'[^\d]', '', cnpj_formatado)
            
            if cnpj_para_consulta and len(cnpj_para_consulta) == 14:
                if cnpj_para_consulta not in cnpjs_unicos:
                    cnpjs_unicos[cnpj_para_consulta] = {
                        'indices': [i],
                        'formatado': cnpj_formatado
                    }
                else:
                    cnpjs_unicos[cnpj_para_consulta]['indices'].append(i)
        
        total_cnpjs = len(cnpjs_unicos)
        print(f"\n📊 Encontrados {total_cnpjs} CNPJ(s) único(s) para consulta")
        
        if total_cnpjs == 0:
            print("Nenhum CNPJ válido encontrado para consulta.")
            return dados
        
        if total_cnpjs > 10:
            tempo_estimado = total_cnpjs * self.cnpj_consultor.delay_seconds / 60
            print(f"⏱️  Tempo estimado: {tempo_estimado:.1f} minutos")
        
        for j, (cnpj, info) in enumerate(cnpjs_unicos.items()):
            indices = info['indices']
            cnpj_formatado_display = info['formatado']
            
            print(f"\n[{j+1}/{total_cnpjs}] Processando: {cnpj_formatado_display}")
            
            dados_cnpj = self.cnpj_consultor.consultar_cnpj(cnpj)
            
            if cnpj in self.cnpj_consultor.cache:
                consultas_cache += 1
            else:
                consultas_realizadas += 1
            
            status_simples = self.cnpj_consultor.get_simples_status(dados_cnpj)
            codigo_cnae, descricao_cnae = self.cnpj_consultor.get_cnae_info(dados_cnpj)
            
            for idx in indices:
                dados[idx]['Consulta Simples API'] = status_simples
                
                regime_xml = dados[idx].get('Simples Nacional / XML', '')
                status_sn = self._status_simples_nacional(regime_xml, status_simples)
                dados[idx]['Status Simples Nacional'] = status_sn
                
                if status_sn == "Divergente":
                    alerta_atual = dados[idx].get('Alertas Fiscais', '')
                    novo_alerta = f"Simples Nacional divergente: XML={regime_xml} vs API={status_simples}"
                    if alerta_atual:
                        dados[idx]['Alertas Fiscais'] = alerta_atual + "; " + novo_alerta
                    else:
                        dados[idx]['Alertas Fiscais'] = novo_alerta
                
                if codigo_cnae and (not dados[idx].get('Código CNAE') or dados[idx].get('Código CNAE') == ''):
                    dados[idx]['Código CNAE'] = codigo_cnae
                if descricao_cnae and (not dados[idx].get('Descrição CNAE') or dados[idx].get('Descrição CNAE') == ''):
                    dados[idx]['Descrição CNAE'] = descricao_cnae
                
                regime_api = status_simples
                codigo_servico = dados[idx].get('Código de serviço', '')
                valor_irrf = self._to_float(dados[idx].get('IRRF', 0))
                valor_inss = self._to_float(dados[idx].get('INSS', 0))
                tipo_retencao_csrf = dados[idx].get('Retenção CSRF', '2 : Sem retenção')
                tem_irrf_retido = valor_irrf > 0
                tem_inss_retido = valor_inss > 0
                
                validacao_api = self.validar_retencoes(dados[idx], regime_api, codigo_servico,
                                                      tipo_retencao_csrf, tem_irrf_retido, tem_inss_retido)
                dados[idx]['_Alertas_API'] = ' | '.join(validacao_api['alertas']) if validacao_api['alertas'] else ''
                
                # INTEGRA OS ALERTAS DA API AO CAMPO PRINCIPAL
                if validacao_api['alertas']:
                    alerta_atual = dados[idx].get('Alertas Fiscais', '')
                    novos_alertas = "; ".join(validacao_api['alertas'])
                    if alerta_atual:
                        dados[idx]['Alertas Fiscais'] = alerta_atual + f"; API: {novos_alertas}"
                    else:
                        dados[idx]['Alertas Fiscais'] = f"API: {novos_alertas}"
            
            print(f"   Resultado Simples Nacional: {status_simples}")
            if codigo_cnae:
                print(f"   CNAE encontrado: {codigo_cnae} - {descricao_cnae[:50]}...")
            
            progresso = (j + 1) / total_cnpjs * 100
            print(f"   Progresso: {progresso:.1f}%")
        
        print(f"\n{'='*60}")
        print("CONSULTA DE SIMPLES NACIONAL E CNAE FINALIZADA")
        print("="*60)
        print(f"✅ Consultas realizadas na API: {consultas_realizadas}")
        print(f"✅ Consultas atendidas por cache: {consultas_cache}")
        print(f"✅ Total de CNPJs processados: {total_cnpjs}")
        
        stats = obter_estatisticas_cache()
        if stats:
            print(f"📊 Cache SQLite: {stats['validos']} registros válidos")
        print("="*60)
        
        print("\n📊 COMPARAÇÃO SIMPLES NACIONAL (XML vs API):")
        for i, item in enumerate(dados[:10]):
            xml_status = item.get('Simples Nacional / XML', 'N/A')
            api_status = item.get('Consulta Simples API', 'N/A')
            status_sn = item.get('Status Simples Nacional', 'N/A')
            cnpj_display = item['CNPJ/CPF'][:20] + "..." if len(item['CNPJ/CPF']) > 20 else item['CNPJ/CPF']
            status_icon = "⚠️" if status_sn == "Divergente" else "✓"
            print(f"  {i+1:3d}. {status_icon} CNPJ: {cnpj_display:<25} XML: {xml_status:<15} API: {api_status:<15} Status: {status_sn}")
        
        if len(dados) > 10:
            print(f"  ... e mais {len(dados) - 10} registro(s)")
        
        # ================================================================
        # RECALCULAR STATUS APÓS CONSULTA API (inclui regra de MEI)
        # ================================================================
        print("\n🔄 Recalculando status de auditoria após consulta API...")
        for item in dados:
            # Regenera o campo _Correcoes para a nova validação
            correcoes = self.verificar_correcoes(item)
            item['_Correcoes'] = correcoes
            
            # Recalcula todos os status (incluindo regra especial de MEI)
            status_sn, status_csrf, status_irrf, status_inss, alertas_fiscais = self.determinar_status_auditoria(item)
            item['Status Simples Nacional'] = status_sn
            item['Status CSRF'] = status_csrf
            item['Status IRRF'] = status_irrf
            item['Status INSS'] = status_inss
            item['Alertas Fiscais'] = alertas_fiscais
        
        print("✅ Status de auditoria recalculados com sucesso!")
        
        return dados
