from __future__ import annotations

import glob
import os
import re
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import date, datetime
from typing import Iterable
from zoneinfo import ZoneInfo

from openpyxl import load_workbook, Workbook

from .nfse_keys import gerar_chave_nfse

SHEET_TODAS = "Todas as Notas"
SHEET_DIVERGENTES = "Notas Divergentes"
SHEET_CORRETAS = "Notas Corretas"

DIA_PROCESSADO_COL = "dia processado"

# Namespaces comumente usados em NFS-e
NAMESPACES = {
    'ns': 'http://www.sped.fazenda.gov.br/nfse',
    'ds': 'http://www.w3.org/2000/09/xmldsig#'
}


def _extrair_chave_acesso_xml(file_path: str) -> str:
    """Extrai a chave de acesso de um arquivo XML.
    
    A chave pode estar:
    1. No atributo 'Id' do elemento infNFSe (ex: NFS21113002250810235000117000000000010426032336425030)
    2. Em elementos como chNFSe, chNfse, chNFS, chNFS-e
    """
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        
        # Tenta encontrar infNFSe
        inf_nfse = root.find('.//ns:infNFSe', NAMESPACES)
        if inf_nfse is None:
            inf_nfse = root.find('.//infNFSe')
        if inf_nfse is None:
            inf_nfse = root.find('.//ns:infDPS', NAMESPACES)
        if inf_nfse is None:
            inf_nfse = root.find('.//infDPS')
        
        if inf_nfse is not None:
            # Primeiro tenta extrair do atributo 'Id'
            chave = inf_nfse.get('Id', '')
            if chave:
                # Remove prefixo comum como "NFS" se presente
                if chave.startswith('NFS'):
                    chave = chave[3:]
                # Remove caracteres não numéricos
                chave = re.sub(r'[^\d]', '', chave)
                if chave:
                    return chave
            
            # Tenta elementos de texto
            for tag in ['chNFSe', 'chNfse', 'chNFS', 'chNFS-e']:
                # Tenta com namespace
                elem = root.find(f'.//ns:{tag}', NAMESPACES)
                if elem is None or elem.text is None:
                    # Tenta sem namespace
                    elem = root.find(f'.//{tag}')
                if elem is not None and elem.text:
                    return elem.text.strip()
        
        return ''
    except Exception as e:
        print(f"  ⚠️ Erro ao ler chave de acesso do XML {file_path}: {e}")
        return ''


def _encontrar_xml_por_documento(xml_dir: str, numero_documento: str) -> str:
    """Encontra o arquivo XML correspondente a um número de documento."""
    if not xml_dir or not numero_documento:
        return ''
    
    # Procura por arquivos XML na pasta
    xml_files = glob.glob(os.path.join(xml_dir, "*.xml"))
    
    for xml_file in xml_files:
        try:
            # Tenta encontrar o número do documento no nome do arquivo
            filename = os.path.basename(xml_file)
            # Remove extensão
            name_without_ext = os.path.splitext(filename)[0]
            # Se o número do documento estiver no nome do arquivo, retorna o caminho
            if numero_documento and str(numero_documento) in name_without_ext:
                return xml_file
        except Exception:
            continue
    
    return ''


def agrupar_por_ano_mes(dados: list[dict]) -> dict[str, list[dict]]:
    """Agrupa registros por AAAA-MM com base na *data da nota*.

    Regra (item 8): usa preferencialmente a data de emissão/processamento da NFS-e (dhProc),
    armazenada em "Data de Emissão" (YYYY-MM-DD). Fallbacks já são tratados no conversor.
    """
    buckets: dict[str, list[dict]] = defaultdict(list)
    for d in dados:
        de = (d.get("Data de Emissão") or "")
        ym = ""
        if isinstance(de, str) and len(de) >= 7 and "-" in de:
            ym = de[:7]
        comp = (d.get("Competência") or "")
        if not ym and isinstance(comp, str) and len(comp) >= 7 and "-" in comp:
            ym = comp[:7]
        if not ym:
            ym = "0000-00"
        buckets[ym].append(d)
    return dict(buckets)


def _ensure_sheet(wb: Workbook, name: str) -> None:
    if name not in wb.sheetnames:
        wb.create_sheet(title=name)


def _get_headers(ws) -> list[str]:
    headers = []
    for cell in ws[1]:
        headers.append((cell.value or "").strip() if isinstance(cell.value, str) else (cell.value or ""))
    # remove trailing empties
    while headers and (headers[-1] == "" or headers[-1] is None):
        headers.pop()
    return [str(h) for h in headers]


def _ensure_header(ws, header: str) -> int:
    headers = _get_headers(ws)
    if header in headers:
        return headers.index(header) + 1
    col = len(headers) + 1
    ws.cell(row=1, column=col, value=header)
    return col


def _index_map(headers: list[str]) -> dict[str, int]:
    return {h: i + 1 for i, h in enumerate(headers)}


def _today_sp() -> str:
    return datetime.now(ZoneInfo("America/Sao_Paulo")).date().isoformat()


def _divergente(d: dict) -> bool:
    # mesma lógica do converter.save_to_excel (para novos itens)
    # Status Base de Cálculo: Se Valor B/C > 0 = Correto, se = 0 E é MEI = Correto, se = 0 E não é MEI = Divergente
    valor_bc = d.get('Valor B/C', 0)
    # Converte para número (trata strings com vírgula, etc.)
    try:
        if isinstance(valor_bc, str):
            valor_bc = valor_bc.replace('.', '').replace(',', '.')
        valor_bc = float(valor_bc)
    except (ValueError, TypeError):
        valor_bc = 0.0
    
    # Se base > 0, não é divergente por base de cálculo
    if valor_bc > 0:
        base_calculo_divergente = False
    else:
        # Verifica se é MEI (XML ou API)
        regime_xml = str(d.get('Simples Nacional / XML', '')).upper()
        regime_api = str(d.get('Consulta Simples API', '')).upper()
        if 'MEI' in regime_xml or 'MEI' in regime_api:
            base_calculo_divergente = False
        else:
            base_calculo_divergente = True
    
    # Verifica outros status
    outros_status = [
        d.get("Status Simples Nacional"),
        d.get("Status CSRF"),
        d.get("Status IRRF"),
        d.get("Status INSS"),
        d.get("Status Valor Líquido")
    ]
    
    has_outros_divergentes = any(
        v is not None and str(v).strip() not in ["Correto", ""]
        for v in outros_status
    )
    
    return base_calculo_divergente or has_outros_divergentes


def _get_key(d: dict) -> str:
    return (d.get("Chave de Acesso") or d.get("chave_nfse") or gerar_chave_nfse(d) or "").strip()


def _read_existing_keys(ws) -> set[str]:
    headers = _get_headers(ws)
    hmap = {h: i for i, h in enumerate(headers)}
    key_col_idx = None
    
    # Tenta encontrar coluna de chave na ordem de prioridade
    for col_name in ["Chave de Acesso", "chave_nfse", "N° Documento"]:
        if col_name in hmap:
            key_col_idx = hmap[col_name] + 1
            break

    keys = set()
    if key_col_idx is None:
        # Se não encontrou nenhuma coluna de chave, retorna conjunto vazio
        # Isso significa que não fará dedupe, adicionará tudo
        return keys

    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=key_col_idx).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            keys.add(s)
    return keys


def _append_rows(ws, headers: list[str], rows: Iterable[dict], dia_proc_col: int, dia_proc_value: str) -> int:
    hmap = _index_map(headers)
    added = 0
    for d in rows:
        next_row = ws.max_row + 1
        for h, col in hmap.items():
            # não escreve dia processado aqui; será sempre a última coluna
            if h == DIA_PROCESSADO_COL:
                continue
            ws.cell(row=next_row, column=col, value=d.get(h, ""))
        # Preenche dia processado APENAS para novas linhas (não sobrescreve valores existentes)
        ws.cell(row=next_row, column=dia_proc_col, value=dia_proc_value)
        added += 1
    return added


def atualizar_planilha_incremental(converter, caminho_planilha: str, novos_dados: list[dict], xml_dir: str = None) -> tuple[int, int]:
    """Atualiza planilha XLSX sem sobrescrever histórico (append-only).

    Regras (item 9):
    - Se existir: NÃO reescreve nenhuma linha existentes (apenas adiciona novas).
    - Dedupe por chave de acesso (quando existir) ou chave_nfse ou N° Documento.
    - Adiciona coluna "dia processado" (última) sem alterar colunas anteriores.
    - Preenche "dia processado" apenas para novas linhas adicionadas.
    - Se xml_dir for fornecido, preenche "Chave de Acesso" para notas existentes que não tiverem.
    """
    if not novos_dados:
        return (0, 0)

    # garante chave_nfse como fallback
    for d in novos_dados:
        if not d.get("chave_nfse") and not d.get("Chave de Acesso"):
            d["chave_nfse"] = gerar_chave_nfse(d)

    os.makedirs(os.path.dirname(caminho_planilha), exist_ok=True)
    dia_proc = _today_sp()

    if not os.path.exists(caminho_planilha):
        # cria do zero com o conversor (mantém layout original) e depois injeta coluna dia processado
        ok = converter.save_to_excel(novos_dados, caminho_planilha)
        if not ok:
            raise RuntimeError("Falha ao criar planilha inicial")
        wb = load_workbook(caminho_planilha)
        for sheet in [SHEET_TODAS, SHEET_DIVERGENTES, SHEET_CORRETAS]:
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            dia_col = _ensure_header(ws, DIA_PROCESSADO_COL)
            # preencher para todas as linhas existentes (todas são novas)
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=dia_col).value in (None, ""):
                    ws.cell(row=r, column=dia_col, value=dia_proc)
        wb.save(caminho_planilha)
        return (0, len(novos_dados))

    # append incremental
    wb = load_workbook(caminho_planilha)
    _ensure_sheet(wb, SHEET_TODAS)
    _ensure_sheet(wb, SHEET_DIVERGENTES)
    _ensure_sheet(wb, SHEET_CORRETAS)

    ws_all = wb[SHEET_TODAS]

    # garante headers na sheet principal
    headers_all = _get_headers(ws_all)
    if not headers_all:
        # planilha vazia: cria header baseado nas chaves do primeiro registro
        base_headers = list(novos_dados[0].keys())
        # assegura chave_nfse no final (antes do dia processado) caso exista
        ws_all.append(base_headers)
        headers_all = base_headers

    # ========================================================================
    # Se xml_dir for fornecido, preenche "Chave de Acesso" para notas existentes
    # ========================================================================
    if xml_dir and os.path.isdir(xml_dir):
        print(f"🔍 Procurando chaves de acesso nos XMLs em: {xml_dir}")
        
        # Encontra o índice da coluna Chave de Acesso
        hmap = {h: i + 1 for i, h in enumerate(headers_all)}
        chave_col_idx = hmap.get("Chave de Acesso")
        
        # Encontra o índice da coluna N° Documento para usar na busca
        doc_col_idx = hmap.get("N° Documento")
        
        if chave_col_idx and doc_col_idx:
            chaves_preenchidas = 0
            for r in range(2, ws_all.max_row + 1):
                # Verifica se a chave de acesso está vazia
                chave_atual = ws_all.cell(row=r, column=chave_col_idx).value
                if not chave_atual or str(chave_atual).strip() == "":
                    # Procura o XML pelo número do documento
                    numero_doc = ws_all.cell(row=r, column=doc_col_idx).value
                    if numero_doc:
                        xml_path = _encontrar_xml_por_documento(xml_dir, str(numero_doc))
                        if xml_path:
                            chave_xml = _extrair_chave_acesso_xml(xml_path)
                            if chave_xml:
                                ws_all.cell(row=r, column=chave_col_idx, value=chave_xml)
                                chaves_preenchidas += 1
                                print(f"  ✓ Chave preenchida para doc {numero_doc}: {chave_xml}")
            
            if chaves_preenchidas > 0:
                print(f"  📊 Total de {chaves_preenchidas} chaves de acesso preenchidas a partir dos XMLs")
                wb.save(caminho_planilha)
                # Recarrega o workbook para atualizar as chaves existentes
                wb = load_workbook(caminho_planilha)
                ws_all = wb[SHEET_TODAS]
                headers_all = _get_headers(ws_all)
    # ========================================================================

    dia_col_all = _ensure_header(ws_all, DIA_PROCESSADO_COL)
    headers_all = _get_headers(ws_all)  # refresh (inclui dia processado se foi criado)
    existing_keys = _read_existing_keys(ws_all)

    to_add_all = []
    to_add_div = []
    to_add_ok = []

    for d in novos_dados:
        k = _get_key(d)
        if not k:
            # Se não tem chave, tenta usar N° Documento como fallback
            k = str(d.get("N° Documento", "")).strip()
        if not k:
            continue
        if k in existing_keys:
            continue
        existing_keys.add(k)
        to_add_all.append(d)
        if _divergente(d):
            to_add_div.append(d)
        else:
            to_add_ok.append(d)

    # garantir headers nas outras sheets compatíveis com a principal
    def sync_headers(ws):
        headers = _get_headers(ws)
        if not headers:
            for h in headers_all:
                ws.cell(row=1, column=headers_all.index(h) + 1, value=h)
            headers = headers_all[:]
        dia_col = _ensure_header(ws, DIA_PROCESSADO_COL)
        return headers, dia_col

    ws_div = wb[SHEET_DIVERGENTES]
    ws_ok = wb[SHEET_CORRETAS]

    headers_div, dia_col_div = sync_headers(ws_div)
    headers_ok, dia_col_ok = sync_headers(ws_ok)

    # append
    added = 0
    added += _append_rows(ws_all, headers_all, to_add_all, dia_col_all, dia_proc)
    _append_rows(ws_div, headers_div, to_add_div, dia_col_div, dia_proc)
    _append_rows(ws_ok, headers_ok, to_add_ok, dia_col_ok, dia_proc)

    wb.save(caminho_planilha)
    return (len(existing_keys) - added, added)

